from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import PatternFill, Color

class InitExcelTemplate:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = '결과보고서'
        self.bucket = 's3inspectortest'

    def createFirstline(self,fileName):
        self.ws['A1'] = '인스턴스'
        cal = self.ws['A1']
        cal.font = Font(bold=True)
        self.ws['B1'] = '카테고리'
        cal = self.ws['B1']
        cal.font = Font(bold=True)
        self.ws['C1'] = '항목번호'
        cal = self.ws['C1']
        cal.font = Font(bold=True)
        self.ws['D1'] = '항목종류'
        cal = self.ws['D1']
        cal.font = Font(bold=True)
        self.ws['E1'] = '항목설명'
        cal = self.ws['E1']
        cal.font = Font(bold=True)
        self.ws['F1'] = '보안대책'
        cal = self.ws['F1']
        cal.font = Font(bold=True)
        self.ws['G1'] = '위험도'
        cal = self.ws['G1']
        cal.font = Font(bold=True)
        self.ws['H1'] = '신규서비스의 경우'
        cal = self.ws['H1']
        cal.font = Font(bold=True)
        self.ws['I1'] = '현재 운영중인 경우'
        cal = self.ws['I1']
        cal.font = Font(bold=True)
        self.ws['J1'] = '조치계획'
        cal = self.ws['J1']
        cal.font = Font(bold=True)
        self.ws['K1'] = '조치내역'
        cal = self.ws['K1']
        cal.font = Font(bold=True)
        self.ws['L1'] = '조치일자'
        cal = self.ws['L1']
        cal.font = Font(bold=True)
        self.ws['M1'] = '이행점검결과'
        cal = self.ws['M1']
        cal.font = Font(bold=True)
        self.wb.save(fileName)
