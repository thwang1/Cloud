import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.styles import PatternFill, Color
from EC2_test import EC2_Controller
from InitExcel import InitExcelTemplate
from ResultFilter import Filter_Controller
import re
import boto3

# self.parsedFindings -> instance -> os,[titleNum] -> title,description, recommendation, etc...
class Report_Controller:
    def __init__(self,detailFindings):
        self.detailFindings = detailFindings
        self.parsedFindings = {}
        self.targetOS = []
        self.targetInstance = []
        self.numFindings = 0
        self.fileNames = []

        for key,value in self.detailFindings.items():
            for ke,val in value.items():
                self.targetInstance.append(val['agentId'])
                self.targetOS.append(val['OSversion'])
        i = 0
        for ins in self.targetInstance:
            temp = {ins: {'OS': self.targetOS[i]}}
            self.parsedFindings.update(temp)
            i = i + 1

    # Inspector JSON 결과를 dict형태로 Parsing
    # 입력 : N/A
    # 출력 : N/A
    def JsonParsing(self):
        cate = ""
        titnum = ""
        for key,value in self.detailFindings.items():
            for ke,val in value.items():
                temp1 = {'accountID': val['accountID'],'vpcID': val['vpcID'],'region': val['region'],'ServiceName': val['ServiceName'],'ManagerName': val['ManagerName'],'region': val['region']}
                k1 = val['agentId']
                self.parsedFindings[k1].update(temp1)
                regex = re.compile(r'^\d+(?:\.\d+)*[ \t]+')
                v = val['id']
                matchobj = regex.search(v)
                num = 0
                if matchobj != None:
                    num = matchobj.group()
                    num = num[:-1]
                    v = v.replace(num,"")
                    temp1 = {num: {'title': v}}
                    self.parsedFindings[k1].update(temp1)
                else:
                    num = v
                    temp1 = {num: {'title': num}}
                    self.parsedFindings[k1].update(temp1)
                z = val['rulesPackageArn']
                if "0-AxKmMHPX" in z or "0-R01qwB5Q" in z or "0-byoQRFYm" in z or "0-JJOtZiqQ" in z or "0-fs0IZZBj" in z or "0-2WRpmi4n" in z or "0-asL6HRgN" in z or "0-bBUQnxMq" in z or "0-ZujVHEPB" in z or "0-SnojL3Z6" in z or "0-XApUiSaP" in z or "0-HfBQSbSf" in z or "0-vlgEGcVD" in z or "0-rOTGqe5G" in z:
                    temp1 = {'category': "Security Best Practices"}
                    self.parsedFindings[k1][num].update(temp1)
                if "0-cE4kTR30" in z or "0-PmNV0Tcd" in z or "0-TxmXimXF" in z or "0-rD1z6dpl" in z or "0-YxKfjFu1" in z or "0-s3OmLzhL" in z or "0-FLcuV4Gz" in z or "0-YI95DVd7" in z or "0-6yunpJ91" in z or "0-SPzU33xe" in z or "0-AizSYyNq" in z or "0-52Sn74uu" in z:
                    temp1 = {'category': "Network Reachability"}
                    self.parsedFindings[k1][num].update(temp1)
                if "0-JnA8Zp85" in z or "0-gEjTy7T7" in z or "0-TKgzoVOa" in z or "0-9hgA516p" in z or "0-LqnJE9dO" in z or "0-PoGHMznc" in z or "0-D5TGAxiR" in z or "0-gHP9oWNT" in z or "0-wNqHa8M9" in z or "0-ubA5XvBh" in z or "0-kZGCqcE1" in z or "0-IgdgIewd" in z or "0-3IFKFuOb" in z or "0-4oQgcI4G" in z:
                    temp1 = {'category': "CVE"}
                    self.parsedFindings[k1][num].update(temp1)
                if "0-m8r61nnh" in z or "0-rExsr2X8" in z or "0-xUY8iRqX" in z or "0-H5hpSawc" in z or "0-PSUlX14m" in z or "0-T9srhg1z" in z or "0-Vkd2Vxjq" in z or "0-7WNjqgGu" in z or "0-nZrAVuv8" in z or "0-sJBhCr0F" in z or "0-IeCjwf1W" in z or "0-Yn8jlX7f" in z or "0-pTLCdIww" in z or "0-Ac4CFOuc" in z:
                    temp1 = {'category': "CIS OS Security Config Benchmarks"}
                    self.parsedFindings[k1][num].update(temp1)
                temp1 = {'severity': val['serverity']}
                self.parsedFindings[k1][num].update(temp1)
                temp1 = {'description': val['description']}
                self.parsedFindings[k1][num].update(temp1)
                temp1 = {'recommendation': val['recommendation']}
                self.parsedFindings[k1][num].update(temp1)

    # 위험도 재조정
    # 입력 : N/A
    # 출력 : N/A
    def changeSeverity(self):
        for instance,value in self.parsedFindings.items():
            osv = value['OS']
            for k,v in value.items():
                if k != 'OS' and k != 'accountID' and k != 'vpcID' and k != 'region' and k != 'ServiceName' and k != 'ManagerName':
                    newSeverity = Filter_Controller(osv).SeverityCheck(osv,v['category'],k)
                    if newSeverity != None:
                        v['severity'] = newSeverity

    # 추가 확인필요 항목 필터링
    # 입력 : N/A
    # 출력 : temp (dict)
    #   - FindingID :
    #           - title
    #           - category
    #           - severity
    #           - description
    #           - recommendation
    def changeAddCheck(self):
        for instance,value in self.parsedFindings.items():
            temp = {}
            os = value['OS']
            for k,v in value.items():
                if k != 'OS' and k != 'accountID' and k != 'vpcID' and k != 'region' and k != 'ServiceName' and k != 'ManagerName':
                    cate = v['category']
                    titNum = k
                    if Filter_Controller(os).AdditionalCheck(os,cate,titNum)==True:
                        temp1 = {titNum: value[titNum]}
                        temp.update(temp1)
            for ke,va in temp.items():
                del(value[ke])
        return temp

    # 신규/운영서비스 관련 코멘트 추가
    # 입력 : N/A
    # 출력 : N/A
    def addNewOrCurrentComment(self):
        for instance,value in self.parsedFindings.items():
            osv = value['OS']
            for k,v in value.items():
                if k != 'OS' and k != 'accountID' and k != 'vpcID' and k != 'region' and k != 'ServiceName' and k != 'ManagerName':
                    cate = v['category']
                    titNum = k
                    comment = Filter_Controller(osv).NewOrCurrentCheck(osv,cate,titNum)
                    if comment != None:
                        temp = {'NewServiceComment': comment[0], 'CurrentServiceComment': comment[1]}
                        self.parsedFindings[instance][titNum].update(temp)

    # 예외처리 항목 필터링
    # 입력 : N/A
    # 출력 : N/A
    def changeNotApplicable(self):
        for instance,value in self.parsedFindings.items():
            temp = {}
            os = value['OS']
            for k,v in value.items():
                if k != 'OS' and k != 'accountID' and k != 'vpcID' and k != 'region' and k != 'ServiceName' and k != 'ManagerName':
                    cate = v['category']
                    titNum = k
                    if Filter_Controller(os).NotApplicableCheck(os,cate,titNum)==True:
                        temp1 = {titNum: value[titNum]}
                        temp.update(temp1)
            for ke,va in temp.items():
                del(value[ke])

    # 보안대책 필터링
    # 입력 : N/A
    # 출력 : N/A
    def changeRemediation(self):
        for instance,value in self.parsedFindings.items():
            osv = value['OS']
            for k,v in value.items():
                if k != 'OS' and k != 'accountID' and k != 'vpcID' and k != 'region' and k != 'ServiceName' and k != 'ManagerName':
                    newRemediation = Filter_Controller(osv).RemediationCheck(osv,v['category'],k)
                    if newRemediation != None:
                        v['recommendation'] = newRemediation


    # 리포팅 엑셀결과파일 생성하는 함수
    # AWS Lambda에서 실행할 경우엔 S3에 템플릿 및 데이터셋 파일 저장 필요
    def DataIntoExcel(self):
        for instance,value in self.parsedFindings.items():
            # empty 엑셀결과파일 생성
            fileName = "FinalReport_" + instance + ".xlsx"
            InitExcelTemplate().createFirstline(fileName)
            wb = openpyxl.load_workbook(fileName)
            ts = wb['결과보고서']

            minrow = 2 # 최소 행 번호
            mincol = 1 # 최소 열 번호
            maxcol = 12 # 최대 열 번호

            # 결과항목 data를 엑셀결과파일에 입력
            row = minrow
            for ke,va in value.items():
                if ke != 'OS' and ke != 'accountID' and ke != 'vpcID' and ke != 'region' and ke != 'ServiceName' and ke != 'ManagerName':
                    if va['severity'] == 'High':
                        for col in range(mincol,maxcol+1):
                            cal = ts.cell(row=row,column=col)
                            if col == 1:
                                cal.value = instance #인스턴스ID
                            if col == 2:
                                cal.value = va['category'] #카테고리(항목종류)
                            if col == 3:
                                cal.value = ke # 항목번호
                            if col == 4:
                                cal.value = va['title'] #항목명
                                self.numFindings = self.numFindings + 1
                            if col == 5:
                                cal.value = va['description'] # 항목설명
                            if col == 6:
                                cal.value = va['recommendation'] # 보안대책
                            if col == 7:
                                cal.value = va['severity'] # 위험도
                            if col == 8:
                                cal.value = va['NewServiceComment'] # 신규서비스의 경우 개발자가 참고할 내용
                                cal.fill = PatternFill(start_color='FFFF63',end_color='FFFF63',fill_type='solid')
                            if col == 9:
                                cal.value = va['CurrentServiceComment'] # 현재 운영중인 경우 개발자가 참고할 내용
                                cal.fill = PatternFill(start_color='FFFF63',end_color='FFFF63',fill_type='solid')
                            if col == 10:
                                cal.value = 'ex) 2020.00.00' # 조치계획
                                cal.font = Font(color="FF0000")
                                cal.fill = PatternFill(start_color='FFBAA4',end_color='FFBAA4',fill_type='solid')
                            if col == 11:
                                cal.value = 'ex) 상세 조치내 기재' # 조치내역
                                cal.font = Font(color="FF0000")
                                cal.fill = PatternFill(start_color='FFBAA4',end_color='FFBAA4',fill_type='solid')
                            if col == 12:
                                ts.cell(row=row,column=col).value = 'ex) 조치완료/조치예정' # 조치일자
                                cal.font = Font(color="FF0000")
                                cal.fill = PatternFill(start_color='FFBAA4',end_color='FFBAA4',fill_type='solid')
                            if col == 13:
                                ts.cell(row=row,column=col).value = '-' # 이행점검결과
                        row = row + 1
            wb.save(fileName)
            self.fileNames.append(fileName)
            print("Excel Final Report (" + fileName + ") Created: Success")
            wb.close()

    # 취약항목 총개수 계산
    def calcNumFindings(self):
        return self.numFindings

    # 생성된 결과엑셀파일 이름 리턴
    def submitFileNames(self):
        return self.fileNames

    def returnParsedFindings(self):
        return self.parsedFindings
